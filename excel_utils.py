import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def guardar_en_excel(fields, ruta_excel):
    # Si existe, leerlo, si no crear DataFrame vacío
    if os.path.exists(ruta_excel):
        df = pd.read_excel(ruta_excel)
    else:
        df = pd.DataFrame(columns=[
            "archivo", "tipo", "numero", "referencia", "asunto",
            "unidad_emisora", "fecha_documento", "firmante"
        ])

    # Actualizar o agregar
    if fields["archivo"] in df["archivo"].values:
        mask = df["archivo"] == fields["archivo"]
        for col in fields:
            df.loc[mask, col] = fields[col]
    else:
        df = pd.concat([df, pd.DataFrame([fields])], ignore_index=True)

    # Guardar Excel con pandas
    df.to_excel(ruta_excel, index=False)

    # --- Ajustar tamaños de columna y wrap text ---
    try:
        wb = load_workbook(ruta_excel)
        ws = wb.active

        col_widths = {
            "A": 35,  # archivo
            "B": 20,  # tipo
            "C": 15,  # numero
            "D": 25,  # referencia
            "E": 50,  # asunto
            "F": 55,  # unidad_emisora
            "G": 25,  # fecha_documento
            "H": 40,  # firmante
        }

        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.column_letter in ["E", "F", "H"]:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        wb.save(ruta_excel)
        wb.close()
    except Exception as e:
        print(f"⚠️ No se pudo ajustar el formato de columnas: {e}")

    #print(f"✅ Datos guardados/actualizados en {ruta_excel}")

