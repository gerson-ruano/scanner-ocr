import os
import re
import pandas as pd
import pytesseract
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pdf2image import convert_from_path

# --- CONFIGURACIÓN ---
# Carpeta donde está el script
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Carpeta "processed" dentro del proyecto
CARPETA_PROCESSED = os.path.join(BASE_DIR, "processed")
CARPETA_ARCHIVOS = os.path.join(BASE_DIR, "invoices")

# Archivo Excel en la carpeta processed
RUTA_EXCEL = os.path.join(CARPETA_PROCESSED, "datos.xlsx")

# Crear carpeta processed si no existe
os.makedirs(CARPETA_PROCESSED, exist_ok=True)

# Si usas Windows y Tesseract no está en PATH, descomenta y pon la ruta correcta:
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
print("Tesseract path:", pytesseract.pytesseract.tesseract_cmd)

# --- FUNCIONES DE OCR ---
def extract_text_with_tesseract(file_path):
    """Extrae texto de imágenes usando Tesseract."""
    try:
        image = Image.open(file_path)
        return pytesseract.image_to_string(image, lang='spa')
    except Exception as e:
        print(f"❌ Error procesando imagen {file_path}: {e}")
        return ""

def extract_text_from_pdf_tesseract(pdf_path):
    try:
        pages = convert_from_path(pdf_path, poppler_path=r"C:\poppler\Library\bin")
        text = ""
        for page in pages:
            text += pytesseract.image_to_string(page, lang='spa')
        return text
    except Exception as e:
        print(f"❌ Error procesando PDF {pdf_path}: {e}")
        return ""

# --- FUNCIONES DE EXTRACCIÓN DE CAMPOS ---
def extract_unidad_emisora(text):
    lines = text.splitlines()
    unidad_emisora = []

    for i, line in enumerate(lines):
        if re.search(r'Ministerio\s+', line, re.IGNORECASE):
            for j in range(1, 4):
                if i + j < len(lines):
                    siguiente_linea = lines[i + j].strip()
                    if re.search(r'OFICIO|Guatemala', siguiente_linea, re.IGNORECASE):
                        break
                    if siguiente_linea:
                        linea_limpia = re.sub(r'^(Ministerio\s+de\s+Gobernación|Gobernación)\s*', '', siguiente_linea, flags=re.IGNORECASE)
                        unidad_emisora.append(linea_limpia)
            break

    return ' '.join(unidad_emisora).strip() if unidad_emisora else None

def extract_fields(text, archivo):
    fields = {}

    # Número de oficio
    match = re.search(r'OFICIO\s+No\.\s+([0-9\-]+)', text, re.IGNORECASE)
    fields['oficio'] = match.group(1) if match else None

    # Referencia
    match = re.search(r'\bREF[./]?\s*[:\-]?\s*([A-Z0-9\-\/]+)', text, re.IGNORECASE)
    fields['referencia'] = match.group(1) if match else None

    # Asunto
    match = re.search(r'ASUNTO\s*:?[\s]*([^\n]+)(?:\n([^\n]+))?', text, re.IGNORECASE)
    if match:
        asunto = match.group(1).strip()
        if match.group(2):
            asunto += " " + match.group(2).strip()
        fields['asunto'] = asunto
    else:
        fields['asunto'] = None

    # Unidad emisora
    fields['unidad_emisora'] = extract_unidad_emisora(text)

    # Firmante
    ultimas_lineas = [l.strip() for l in text.strip().split('\n') if l.strip() != ''][-20:]
    firmante = None
    for i in range(len(ultimas_lineas) - 1, 1, -1):
        if re.search(r'(LIC\.|OFICIAL|JEFE|DIRECTOR|COORDINADOR|ENCARGADO)', ultimas_lineas[i], re.IGNORECASE):
            linea1 = ultimas_lineas[i - 1]
            linea2 = ultimas_lineas[i - 2] if i >= 2 else ''
            firmante = f"{linea2} / {linea1} / {ultimas_lineas[i]}".strip(' /')
            break
        fields['firmante'] = firmante

    # Fecha del documento
    match = re.search(r'Guatemala,\s+(\d{1,2}\s+de\s+\w+\s+de\s+\d{4})', text)
    fields['fecha_documento'] = match.group(1) if match else None

    # Nombre de archivo
    fields['archivo'] = archivo

    return fields

# --- GUARDAR EN EXCEL ---
def guardar_en_excel(fields, ruta_excel=RUTA_EXCEL):
    # Si existe, leerlo, si no crear DataFrame vacío
    if os.path.exists(ruta_excel):
        df = pd.read_excel(ruta_excel)
    else:
        df = pd.DataFrame(columns=["archivo", "oficio", "referencia", "asunto", "unidad_emisora", "firmante", "fecha_documento"])

    # Actualizar o agregar
    if fields["archivo"] in df["archivo"].values:
        mask = df["archivo"] == fields["archivo"]
        for col in fields:
            df.loc[mask, col] = fields[col]
    else:
        df = pd.concat([df, pd.DataFrame([fields])], ignore_index=True)

    # Guardar Excel con pandas
    df.to_excel(ruta_excel, index=False)

    # --- Ajustar tamaños de columna y wrap text ---
    try:
        wb = load_workbook(ruta_excel)
        ws = wb.active

        # Definir anchos personalizados
        col_widths = {
            "A": 30,  # archivo
            "B": 15,  # oficio
            "C": 20,  # referencia
            "D": 50,  # asunto
            "E": 55,  # unidad_emisora
            "F": 40,  # firmante
            "G": 20,  # fecha_documento
        }

        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # Aplicar ajuste de texto en columnas largas
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.column_letter in ["D", "E", "F"]:  # columnas de texto largo
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        wb.save(ruta_excel)
        wb.close()
    except Exception as e:
        print(f"⚠️ No se pudo ajustar el formato de columnas: {e}")

    print(f"✅ Datos guardados/actualizados en {ruta_excel}")

# --- PROCESAR TODOS LOS ARCHIVOS ---
def main():
    if not os.path.exists(CARPETA_ARCHIVOS):
        print(f"❌ Carpeta no encontrada: {CARPETA_ARCHIVOS}")
        return

    for nombre_archivo in os.listdir(CARPETA_ARCHIVOS):
        ruta_completa = os.path.join(CARPETA_ARCHIVOS, nombre_archivo)

        if not os.path.isfile(ruta_completa):
            continue

        if nombre_archivo.lower().endswith('.pdf'):
            texto = extract_text_from_pdf_tesseract(ruta_completa)
        elif nombre_archivo.lower().endswith(('.jpg', '.jpeg', '.png', '.tiff')):
            texto = extract_text_with_tesseract(ruta_completa)
        else:
            continue

        print(f"📄 Procesando: {nombre_archivo}")

        if texto.strip():
            campos = extract_fields(texto, nombre_archivo)
            guardar_en_excel(campos)
        else:
            print(f"⚠️ No se pudo extraer texto de {nombre_archivo}")

    print("✅ Proceso completado. Revisa el archivo Excel.")

# --- EJECUCIÓN ---
if __name__ == "__main__":
    main()
